import openpyxl, os
from time import time

# def find_correct_last_row(in_worksheet):
#     """
#     get sheet
#     return last row number
#     """
#     last_filled_row = 0
#     for index, row in enumerate(in_worksheet.values, start=1):
#         corrent_row_empty = row.count(None) == len(row)
#         if corrent_row_empty:
#             empty_row_successively += 1
#         else:
#             empty_row_successively = 0
#             last_filled_row = index
#         if empty_row_successively >= 5:
#             break
#     return last_filled_row

# file = 'school3.xlsx'
# workbook =  openpyxl.load_workbook(file, keep_vba=False, read_only=True)
# demension = workbook['5. Сведения о кадрах'].calculate_dimension()
# print(demension)
# workbook['5. Сведения о кадрах'].reset_dimensions()
# demension = workbook['5. Сведения о кадрах'].calculate_dimension()
# print(demension)
# worksheet = workbook['5. Сведения о кадрах']
# #print(worksheet.max_row)
# for iteration_worksheet in workbook.worksheets:
#     print("Current: ", iteration_worksheet.max_row, "Fact: ", find_correct_last_row(iteration_worksheet))
# workbook.close()

# write_workbook = openpyxl.load_workbook(file, read_only=True)
# write_worksheet = write_workbook['2. Сведения об обучающихся']
#new_workbook = openpyxl.Workbook(write_only=True)
#new_workbook.create_sheet(title="copy_sheet")
#new_workbook.save(time(),'.xlsx')
#help(write_worksheet)
#help(write_worksheet.values)
#help(write_worksheet.cell(2,2).style_array)
#help(write_worksheet.iter_rows)
# for item in write_worksheet.iter_rows(3):
#     print(item)
#     #help(item)
#     break
# write_workbook.close() 

# def remove_exscess_row(workbook):
#     for iteration_worksheet in workbook.worksheets:
#         first_empty_row = find_correct_last_row(iteration_worksheet) + 1
#         max_row = iteration_worksheet.max_row
#         iteration_worksheet.delete_rows(first_empty_row, max_row-first_empty_row+1)


# wb = openpyxl.load_workbook('school3.xlsx')
# remove_exscess_row(wb)
# wb.save('school3.xlsx')
# wb.close()

file_list = ['C:\\Users\\alist\\OneDrive\\Документы\\osokin\\monitoring\\example.xlsx', 
             'C:\\Users\\alist\\OneDrive\\Документы\\osokin\\monitoring\\school3.xlsx', 
             'C:\\Users\\alist\\OneDrive\\Документы\\osokin\\monitoring\\babaevo\\school1.xlsx', 
             'C:\\Users\\alist\\OneDrive\\Документы\\osokin\\monitoring\\babaevo\\school2.xlsx', 
             'C:\\Users\\alist\\OneDrive\\Документы\\osokin\\monitoring\\babaevo\\school3.xlsx', 
             'C:\\Users\\alist\\OneDrive\\Документы\\osokin\\monitoring\\chagoda\\school2.xlsx']

new_list  = [os.path.split(file) for file in file_list]
print(new_list)
print(os.path.join(new_list[0][0], new_list[0][1]))
#print(os.path.abspath(file_list[0]))
#print(os.path.basename(file_list[0]))
#print(os.path.commonpath(file_list[0]))
#print(os.path.dirname(file_list[0]))

def dialog_menu(value):
    """
    """
def get_directory_sorted_list(dir):
    """
    Take list of currect file
    return sorted list dirctory
    """
    path_list = sorted(list({ os.path.dirname(file) for file in dir }), key=len, reverse=True)
    return path_list
def get_current_file_from_directory(file_list, dir_name):
    """
    return all needed file to work
    """
    work_file_list = list()
    for file in file_list:
        if os.path.split(file)[0] == dir_name:
            work_file_list.append(file)
    return work_file_list

def create_needed_sheet(workbook_name, sheet_list):
    """
    Create page if then not have
    """
    book = openpyxl.load_workbook(workbook)
    for sheet_name in sheet_list:
        if not hasattr(book, sheet_name):
            book.create_sheet(sheet_name)
    del(book['Sheet'])
    book.save(workbook)
    book.close()
def create_needed_sheets(workbook_list, sheet_list):
    """
    """
    pass
def file_generator(file_list):
    """
    Create file by list
    """
    created_file = list()
    for file in file_list:
        if not os.path.exists(f'{file}.xlsx'):
            new_book = openpyxl.Workbook()
            new_book.save(f'{file}.xlsx')
            new_book.close()
            created_file.append(f'{file}.xlsx')
    return created_file
def create_sheet_content(destination_file, source_files):
    """
    """
    dest_file = openpyxl.load_workbook(destination_file)
    exam_file = openpyxl.load_workbook('example.xlsx')
    for work_sheet in dest_file.worksheets:
        for row in exam_file[work_sheet.title].rows:
            work_sheet.append([add_data_to_cell(cell, source_files) for cell in row])
    dest_file.save(destination_file)
    dest_file.close()
    exam_file.close()

def make_sheet_style(destination_file):
    """
    """
    dest_file = openpyxl.load_workbook(destination_file)
    exam_file = openpyxl.load_workbook('example.xlsx')
    for work_sheet in dest_file.worksheets:
        for row in exam_file[work_sheet.title].rows:
            for cell in row:
                if cell.has_style:
                    pass
    dest_file.save(destination_file)
    dest_file.close()
    exam_file.close()

def add_data_to_cell(cell, source_files):
    """
    """
    data_cell = None
    if isinstance(cell.value, str):
        data_cell = cell.value
    else:
        data_cell = "="
        page_title = cell.parent.title
        for address in source_files:
            if data_cell[-1] != "=":
                data_cell += "+"
            data_cell += "\'"
            data_cell += os.path.dirname(address)
            data_cell += os.sep
            data_cell += "["
            data_cell += os.path.basename(address)
            data_cell += "]"
            data_cell += page_title
            data_cell += "\'"
            data_cell += "!"
            data_cell += cell.coordinate
    return data_cell

dir_list = get_directory_sorted_list(file_list)
example_file = openpyxl.load_workbook(filename="example.xlsx", read_only=True)
example_file.close()
sheet_names = example_file.sheetnames
created_file = file_generator(dir_list)
for workbook in created_file:
    create_needed_sheet(workbook,sheet_names)
    directory_needed_files = get_current_file_from_directory(file_list, os.path.abspath(workbook[:-5]))
#    create_sheet_content(workbook, directory_needed_files)



input('Press Enter to Continue...')